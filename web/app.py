from flask import Flask, render_template, request, redirect, url_for, jsonify,make_response,session,send_file
from flask.sessions import NullSession
from flask_socketio import SocketIO
from flask_login.utils import logout_user
import pandas as pd
import pyodbc
import pprint
import pdfkit
import json
from datetime import datetime, timedelta
import decimal
import time 
from openpyxl.styles import Alignment , Font , PatternFill
import openpyxl

class Encoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, decimal.Decimal): return float(obj)
class create_dict(dict): 
  
    # __init__ function 
    def __init__(self): 
        self = dict() 
          
    # Function to add key:value 
    def add(self, key, value): 
        self[key] = value


#ddddd
app = Flask(__name__)

count = 3
@app.route('/')
@app.route('/batch_report')
def batch_report():
    #cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ host +';DATABASE='+database+';UID='+user+';PWD='+passwd)
    #SQL_Login = cnxn.cursor()
    #SQL_Login.execute("Exec USP_Journal_Rpt '" + Campaign_ID + "','"+ Lot_ID +"','"+ StartOfPeriod_DT + "','"+EndOfPeriod_DT+"'" )
    return render_template('batch_report.html')

@app.route('/report')
def report():
    #cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ host +';DATABASE='+database+';UID='+user+';PWD='+passwd)
    #SQL_Login = cnxn.cursor()
    #SQL_Login.execute("Exec USP_Journal_Rpt '" + Campaign_ID + "','"+ Lot_ID +"','"+ StartOfPeriod_DT + "','"+EndOfPeriod_DT+"'" )
    
    return render_template('report.html')


@app.route('/report/RawMaterial/<string:pdOrder>')
def Raw_Material(pdOrder):
    return render_template('reportRawMaterial.html',pdOrder=pdOrder)

@app.route('/report/Raw_Material_API/<string:pdOrder>')
def Raw_Material_API(pdOrder):
    host = "172.30.1.1"
    port = 1433
    database = "managedb"
    user = "sa"
    passwd = "qwerty@2019"
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ host +';DATABASE='+database+';UID='+user+';PWD='+passwd)
    RM = cnxn.cursor()
    RM.execute("SELECT * FROM [dbo].[View_RM_Process_tab] where RM_PD_ORDER = '" + pdOrder + "'")
    
    payload = []
    content = {}
    for result in RM:
        
        content = {'RM_RAW_ID': result[1], 'RM_NAME': str(result[2]), 'RM_UNIT': result[3],'PW_NEO_NAME': result[4],'RM_PW_TARGET': result[5],'RM_PW_ACTUAL': str(result[6]),'RM_REQ_DATE': str(result[7]),'RM_INS_DT': str(result[8]),'PW_ITEM_TOTAL': str(result[9]),'RM_PW_EACH_ITEM': str(result[10]),'RM_BARCODE_LIST': str(result[11]),'RM_FFTANK_STATUS': str(result[12]),'RM_FFTANK_1': str(result[13]),'RM_SFTANK_1': str(result[14]),'RM_REF_NO': str(result[15]),'UserDisplayName': str(result[16])}
        payload.append(content)
        content = {}
    #print(payload)
    return json.dumps({"data":payload}, cls = Encoder), 201

@app.route('/report/Overview/<string:pdOrder>')
def reportOverview(pdOrder):    
    
    return render_template('reportOverview.html',pdOrder=pdOrder)


@app.route('/report/PreMixing/<string:pdOrder>')
def reportPreMixing(pdOrder):    
    server = "172.30.2.2"
    port = 5432
    database = "OEE_DB"
    username = "sa"
    password = "p@ssw0rd"
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    PreMixing = cnxn.cursor()
    PreMixing.execute("SELECT RecordID, PD_ORDER, PhaseID, Status, Start_time, End_Time, SetPoint1, Actual1, SetPoint2, Actual2, SetPoint3, Actual3, SetPoint4, Actual4, SetPoint5, Actual5, SetPoint6, Actual6, SetPoint7, Actual7, SetPoint8, Actual8, User_Mixing ,DATEDIFF(second,Start_time,End_Time) as Time_Sec  FROM SCADA_DB.dbo.Mixing_Report mr  WHERE PhaseID > 100 AND PhaseID < 200 AND PD_ORDER = '" +pdOrder+ "' ORDER BY Start_time ASC")
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    Phase_Parameter = cnxn.cursor()
    Phase_Parameter.execute("SELECT PhaseID ,PhaseName,Parameter1,Parameter2,Parameter3,Parameter4,Parameter5,Parameter6,Parameter7,Parameter8 FROM SCADA_DB.dbo.Phase_Parameter") 
    
    Phase_Parameter_DIR = Phase_Parameter.fetchall()

    insertObject = []
    columnNames = [column[0] for column in Phase_Parameter.description]
    for record in Phase_Parameter_DIR:
        insertObject.append( dict( zip( columnNames , record ) ) )
    print(insertObject)
    
    return render_template('reportPreMixing.html',pdOrder=pdOrder,Phase_Parameter=insertObject,PreMixing=PreMixing,len=len(Phase_Parameter_DIR))

@app.route('/report/MainMixing/<string:pdOrder>')
def reportMainMixing(pdOrder):    
    server = "172.30.2.2"
    port = 5432
    database = "OEE_DB"
    username = "sa"
    password = "p@ssw0rd"
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    MainMixing = cnxn.cursor()
    MainMixing.execute("SELECT RecordID, PD_ORDER, PhaseID, Status, Start_time, End_Time, SetPoint1, Actual1, SetPoint2, Actual2, SetPoint3, Actual3, SetPoint4, Actual4, SetPoint5, Actual5, SetPoint6, Actual6, SetPoint7, Actual7, SetPoint8, Actual8, User_Mixing ,DATEDIFF(second,Start_time,End_Time) as Time_Sec  FROM SCADA_DB.dbo.Mixing_Report mr  WHERE PhaseID > 1 AND PhaseID < 100 AND PD_ORDER = '" +pdOrder+ "' ORDER BY Start_time ASC")
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    Phase_Parameter = cnxn.cursor()
    Phase_Parameter.execute("SELECT PhaseID ,PhaseName,Parameter1,Parameter2,Parameter3,Parameter4,Parameter5,Parameter6,Parameter7,Parameter8 FROM SCADA_DB.dbo.Phase_Parameter") 
    
    Phase_Parameter_DIR = Phase_Parameter.fetchall()

    insertObject = []
    columnNames = [column[0] for column in Phase_Parameter.description]
    for record in Phase_Parameter_DIR:
        insertObject.append( dict( zip( columnNames , record ) ) )
    print(insertObject)
    
    return render_template('reportMainMixing.html',pdOrder=pdOrder,Phase_Parameter=insertObject,MainMixing=MainMixing,len=len(Phase_Parameter_DIR))


@app.route('/report/MixingStorage/<string:pdOrder>')
def reportMixingStorage(pdOrder):
    server = "172.30.2.2"
    port = 5432
    database = "OEE_DB"
    username = "sa"
    password = "p@ssw0rd"
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    MixingStorage = cnxn.cursor()
    MixingStorage.execute("SELECT RecordID, PD_ORDER, PhaseID, Status, Start_time, End_Time, SetPoint1, Actual1, SetPoint2, Actual2, SetPoint3, Actual3, SetPoint4, Actual4, SetPoint5, Actual5, SetPoint6, Actual6, SetPoint7, Actual7, SetPoint8, Actual8, User_Mixing ,DATEDIFF(second,Start_time,End_Time) as Time_Sec  FROM SCADA_DB.dbo.Mixing_Report mr  WHERE PhaseID > 200 AND PhaseID < 300 AND PD_ORDER = '" +pdOrder+ "' ORDER BY Start_time ASC")
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    Phase_Parameter = cnxn.cursor()
    Phase_Parameter.execute("SELECT PhaseID ,PhaseName,Parameter1,Parameter2,Parameter3,Parameter4,Parameter5,Parameter6,Parameter7,Parameter8 FROM SCADA_DB.dbo.Phase_Parameter") 
    
    Phase_Parameter_DIR = Phase_Parameter.fetchall()

    insertObject = []
    columnNames = [column[0] for column in Phase_Parameter.description]
    for record in Phase_Parameter_DIR:
        insertObject.append( dict( zip( columnNames , record ) ) )
    print(insertObject)
    
    return render_template('reportMixingStorage.html',pdOrder=pdOrder,Phase_Parameter=insertObject,MixingStorage=MixingStorage,len=len(Phase_Parameter_DIR))
 
@app.route('/report/SidePOT_1/<string:pdOrder>')
def reportSidePOT_1(pdOrder):    
    server = "172.30.2.2"
    port = 5432
    database = "OEE_DB"
    username = "sa"
    password = "p@ssw0rd"
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    SidePOT_1 = cnxn.cursor()
    SidePOT_1.execute("SELECT RecordID, PD_ORDER, PhaseID, Status, Start_time, End_Time, SetPoint1, Actual1, SetPoint2, Actual2, SetPoint3, Actual3, SetPoint4, Actual4, SetPoint5, Actual5, SetPoint6, Actual6, SetPoint7, Actual7, SetPoint8, Actual8, User_Mixing ,DATEDIFF(second,Start_time,End_Time) as Time_Sec  FROM SCADA_DB.dbo.Mixing_Report mr  WHERE PhaseID > 300 AND PhaseID < 400 AND PD_ORDER = '" +pdOrder+ "' ORDER BY Start_time ASC")
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    Phase_Parameter = cnxn.cursor()
    Phase_Parameter.execute("SELECT PhaseID ,PhaseName,Parameter1,Parameter2,Parameter3,Parameter4,Parameter5,Parameter6,Parameter7,Parameter8 FROM SCADA_DB.dbo.Phase_Parameter") 
    
    Phase_Parameter_DIR = Phase_Parameter.fetchall()

    insertObject = []
    columnNames = [column[0] for column in Phase_Parameter.description]
    for record in Phase_Parameter_DIR:
        insertObject.append( dict( zip( columnNames , record ) ) )
    print(insertObject)
    
    return render_template('reportSidePOT_1.html',pdOrder=pdOrder,Phase_Parameter=insertObject,SidePOT_1=SidePOT_1,len=len(Phase_Parameter_DIR))

@app.route('/report/SidePOT_2/<string:pdOrder>')
def reportSidePOT_2(pdOrder):    
    server = "172.30.2.2"
    port = 5432
    database = "OEE_DB"
    username = "sa"
    password = "p@ssw0rd"
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    SidePOT_2 = cnxn.cursor()
    SidePOT_2.execute("SELECT RecordID, PD_ORDER, PhaseID, Status, Start_time, End_Time, SetPoint1, Actual1, SetPoint2, Actual2, SetPoint3, Actual3, SetPoint4, Actual4, SetPoint5, Actual5, SetPoint6, Actual6, SetPoint7, Actual7, SetPoint8, Actual8, User_Mixing ,DATEDIFF(second,Start_time,End_Time) as Time_Sec  FROM SCADA_DB.dbo.Mixing_Report mr  WHERE PhaseID > 400 AND PhaseID < 500 AND PD_ORDER = '" +pdOrder+ "' ORDER BY Start_time ASC")
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    Phase_Parameter = cnxn.cursor()
    Phase_Parameter.execute("SELECT PhaseID ,PhaseName,Parameter1,Parameter2,Parameter3,Parameter4,Parameter5,Parameter6,Parameter7,Parameter8 FROM SCADA_DB.dbo.Phase_Parameter") 
    
    Phase_Parameter_DIR = Phase_Parameter.fetchall()

    insertObject = []
    columnNames = [column[0] for column in Phase_Parameter.description]
    for record in Phase_Parameter_DIR:
        insertObject.append( dict( zip( columnNames , record ) ) )
    print(insertObject)
    
    return render_template('reportSidePOT_2.html',pdOrder=pdOrder,Phase_Parameter=insertObject,SidePOT_2=SidePOT_2,len=len(Phase_Parameter_DIR))

@app.route('/report/QC/<string:pdOrder>')
def reportQC(pdOrder):    
    server = "172.30.2.2"
    database = "OEE_DB"
    username = "sa"
    password = "p@ssw0rd"
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    PD_QC = cnxn.cursor()
    PD_QC.execute("SELECT * FROM SCADA_DB.dbo.QC_Process WHERE PD_Order = ?",(pdOrder,))
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    PD_QC2 = cnxn.cursor()
    PD_QC2.execute("SELECT * FROM SCADA_DB.dbo.QC_Process WHERE PD_Order = ?",(pdOrder,))
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    PD_QC3 = cnxn.cursor()
    PD_QC3.execute("SELECT * FROM SCADA_DB.dbo.QC_Process WHERE PD_Order = ?",(pdOrder,))
    dataPD = []
    dataPD_len = 0
    for i in PD_QC3:
       
        dataPD.append(i[9])
    dataPD_len = len(dataPD) - 1
    return render_template('reportQC.html',pdOrder=pdOrder,PD_QC = PD_QC,PD_QC2=PD_QC2,dataPD = dataPD,dataPD_len=dataPD_len)


@app.route('/batch_report_API' ,methods=["GET", "POST"])
def batch_report_API():
    global count
    q = request.args.get('q')
    
    print(q)

    if request.method == "POST":
        count+=1
        return redirect(url_for('ReportOEE'))
    else:
        host = "172.30.1.1"
        port = 1433
        database = "managedb"
        user = "sa"
        passwd = "qwerty@2019"
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ host +';DATABASE='+database+';UID='+user+';PWD='+passwd)
        batch_report = cnxn.cursor()
        batch_report.execute("SELECT PD_ORDER , PD_PLAN_DT , PD_TARGET_QTY ,PD_UNIT , PD_FM_CODE,PD_FM_NAME , PD_BATCHNO ,PD_PROC_P_ST , PD_PROC_O_ST , PD_PROC_M_ST , PD_PROC_Q_ED , PD_PROC_S_DT ,PD_STATUS_CODE FROM [dbo].[PD_ORDER_VIEW_RPT] ")
        
    
        payload = []
        content = {}
        for result in batch_report:
            content = {'PD': result[0], 'PD_PLAN_DT': str(result[1]), 'PD_TARGET_QTY': result[2],'PD_UNIT': result[3],'PD_FM_CODE': result[4]+'/'+result[5],'PD_BATCHNO': result[6],'PD_PROC_P_ST': str(result[7]),'PD_PROC_O_ST': str(result[8]),'PD_PROC_M_ST': str(result[9]),'PD_PROC_Q_ED': str(result[10]),'PD_PROC_S_DT': str(result[11]),'PD_STATUS_CODE': str(result[12])}
            payload.append(content)
            content = {}
        #print(payload)
        return json.dumps({"data":payload}, cls = Encoder), 201
         
@app.route('/QC_report')
def QC_report():    
    return render_template('QC_report.html') 


@app.route('/Report_QC_Excel')
def Report_QC_Excel():
  
    df = pd.read_json('http://192.168.1.145:5001//QC_report_Excel_API')
    print(df)
    df.to_excel('QC_report.xlsx',index=False)
    
    time.sleep(1)
    
    wabu = openpyxl.load_workbook('QC_report.xlsx')
    washi = wabu.active
    washi.insert_rows(1,5)
    washi['A1'] = 'QC Report'
    washi['A1'].alignment = Alignment(vertical='center')
    washi['A1'].alignment = Alignment(horizontal='center')
    washi['A1'].font = Font(color='FFFFFF',
                        size=24,bold=True)
    washi['A1'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi.merge_cells('A1:F1')
    washi.column_dimensions['A'].width = 20
    washi.column_dimensions['B'].width = 20
    washi.column_dimensions['C'].width = 20
    washi.column_dimensions['D'].width = 20
    washi.column_dimensions['E'].width = 20
    washi.column_dimensions['F'].width = 20
    washi.column_dimensions['G'].width = 20
    washi.column_dimensions['H'].width = 20
    
   
    wabu.save('QC_report.xlsx')
    
    return send_file('..\QC_report.xlsx') 

#ss0000

@app.route('/QC_report_Excel_API' ,methods=["GET", "POST"])
def QC_report_Excel_API():

    
    server = "172.30.2.2"
    port = 5432
    database = "OEE_DB"
    username = "sa"
    password = "p@ssw0rd"
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    PD_QC = cnxn.cursor()
    PD_QC.execute("SELECT DISTINCT PD_Order FROM SCADA_DB.dbo.QC_Process")
    payload = []
    for i in  PD_QC:
        print(i[0])
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
        qc_report = cnxn.cursor()
        qc_report.execute("""
                          SELECT TOP(1) Product_Name  , PD_Order , [Lot_No.] ,BAY ,[Tank_S/N]   , (SELECT TOP(1) [DateTime] 
                            FROM SCADA_DB.dbo.QC_Process WHERE [Action] = 'QC_RECEIVE' AND  PD_Order = ? ) As QC_START 
                            , (SELECT TOP(1) [DateTime] FROM SCADA_DB.dbo.QC_Process WHERE ([Action] = 'QC_PASS' OR [Action] = 'QC_REJECT') AND PD_Order = ? ) As QC_FINISH
                            , DATEDIFF(MINUTE , (SELECT TOP(1) [DateTime] FROM SCADA_DB.dbo.QC_Process WHERE [Action] = 'QC_RECEIVE' AND  PD_Order = ? ) ,
                            (SELECT TOP(1) [DateTime] FROM SCADA_DB.dbo.QC_Process WHERE ([Action] = 'QC_PASS' OR [Action] = 'QC_REJECT') AND PD_Order = ? )) ,
                            [User] 
                            FROM SCADA_DB.dbo.QC_Process WHERE PD_Order = ? 
                          """,(i[0],i[0],i[0],i[0],i[0]))
       

        
        content = {}
        for result in qc_report:
            content = {'Product_Name': str(result[0]), 'PD_Order': result[1],'Lot_No': result[2],'BAY': result[3],'Tank_SN': result[4],'QC_START': str(result[5]),'QC_FINISH': str(result[6]),'QC_TIME': str(result[7]),'User': str(result[8])}
            payload.append(content)
            content = {}
    return json.dumps(payload, cls = Encoder), 201

@app.route('/QC_report_API',methods=["GET", "POST"])
def QC_report_API():    
 
    server = "172.30.2.2"
    port = 5432
    database = "OEE_DB"
    username = "sa"
    password = "p@ssw0rd"
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
    PD_QC = cnxn.cursor()
    PD_QC.execute("SELECT DISTINCT PD_Order FROM SCADA_DB.dbo.QC_Process")
    payload = []
    for i in  PD_QC:
        print(i[0])
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+ server +';DATABASE='+database+';UID='+username+';PWD='+password)
        qc_report = cnxn.cursor()
        qc_report.execute("""
                          SELECT TOP(1) Product_Name  , PD_Order , [Lot_No.] ,BAY ,[Tank_S/N] ,(SELECT COUNT(*) FROM SCADA_DB.dbo.QC_Process WHERE [Action] = 'QC_RECEIVE' AND PD_Order = ?) 
                          , (SELECT TOP(1) [DateTime] FROM SCADA_DB.dbo.QC_Process WHERE [Action] = 'QC_RECEIVE' AND  PD_Order = ? ) As QC_START 
                            , (SELECT TOP(1) [DateTime] FROM SCADA_DB.dbo.QC_Process WHERE ([Action] = 'QC_PASS' OR [Action] = 'QC_REJECT') AND PD_Order = ? ) As QC_FINISH
                            , DATEDIFF(MINUTE , (SELECT TOP(1) [DateTime] FROM SCADA_DB.dbo.QC_Process WHERE [Action] = 'QC_RECEIVE' AND  PD_Order = ? ) ,
                            (SELECT TOP(1) [DateTime] FROM SCADA_DB.dbo.QC_Process WHERE ([Action] = 'QC_PASS' OR [Action] = 'QC_REJECT') AND PD_Order = ? )) ,
                            [User] 
                            FROM SCADA_DB.dbo.QC_Process WHERE PD_Order = ? 
                          """,(i[0],i[0],i[0],i[0],i[0],i[0]))
       

        
        content = {}
        for result in qc_report:
            content = {'Product_Name': str(result[0]), 'PD_Order': result[1],'Lot_No': result[2],'BAY': result[3],'Tank_SN': result[4],'NO': result[5],'QC_START': str(result[6]),'QC_FINISH': str(result[7]),'QC_TIME': str(result[8]),'User': str(result[9])}
            payload.append(content)
            content = {}
    #print(payload)
    return json.dumps({"data":payload}, cls = Encoder), 201

        
if __name__ == "__main__":
    app.run(host='0.0.0.0', debug=True ,port=5001)
